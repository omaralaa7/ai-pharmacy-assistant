"""
Phase 1 — Excel Data Processing & Knowledge Base Construction
AI-Driven Insurance Chatbot (Diploma Project)

Reads the real Egyptian insurance dispensing rules Excel file and converts it
into a structured JSON knowledge base that the chatbot engine can search.

Input:  نظم صرف شركات التأمين-2.xlsx (777 rows, 77 companies, 14 categories)
Output: data/insurance_knowledge_base.json
"""

import os
import re
import sys
import json
import openpyxl

# Fix Windows console encoding for Arabic output
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(os.path.dirname(BASE_DIR), "نظم صرف شركات التأمين-2.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "insurance_knowledge_base.json")

# ---------------------------------------------------------------------------
# Category name translations (for bilingual search support)
# ---------------------------------------------------------------------------
CATEGORY_TRANSLATIONS = {
    "نماذج الصرف": "Dispensing Forms",
    "التحمل": "Co-payment / Deductible",
    "التشخيص": "Diagnosis Requirements",
    "صلاحية النموذج": "Form Validity Period",
    "صورة البطاقة": "National ID Copy Requirement",
    "صورة الكارنية": "Insurance Card Copy Requirement",
    "الختم / إمضاء العميل": "Stamp / Patient Signature Requirements",
    "المحظورات": "Excluded / Prohibited Items",
    "أقصى مدة للصرف": "Maximum Dispensing Duration",
    "الحد الأقصى": "Maximum Financial Limit",
    "التواصل للموافقات": "Approval Contact Information",
    "لينك الاونلاين سيستم": "Online System Link",
    "البدائل": "Generic Substitution Rules",
    "ملاحظات": "Additional Notes",
}


def clean_text(text):
    """Clean and normalize text from the Excel file."""
    if not text or text == "None":
        return ""
    text = str(text).strip()
    # Remove non-breaking spaces
    text = text.replace("\xa0", " ")
    # Normalize multiple newlines
    text = re.sub(r"\n{3,}", "\n\n", text)
    # Remove trailing whitespace per line
    text = "\n".join(line.rstrip() for line in text.split("\n"))
    return text.strip()


def extract_company_names(raw_name):
    """
    Extract Arabic and English names from the company field.

    Examples:
        "يونايتد-united" -> ("يونايتد", "united")
        "GLOBEMED-جلوبميد" -> ("جلوبميد", "GLOBEMED")
        "الاكاديمية البحرية للعلوم" -> ("الاكاديمية البحرية للعلوم", "")
    """
    if not raw_name:
        return "", ""

    raw_name = clean_text(raw_name)

    # Split on hyphen and classify by character ranges
    parts = re.split(r"[-–]", raw_name)
    arabic = ""
    english = ""

    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Check if part contains Arabic characters
        has_arabic = bool(re.search(r"[\u0600-\u06FF]", part))
        has_latin = bool(re.search(r"[a-zA-Z]", part))

        if has_arabic and not arabic:
            arabic = part
        elif has_latin and not english:
            english = part
        elif has_arabic:
            arabic = arabic + "-" + part if arabic else part

    # If no split worked, assign the whole name to Arabic
    if not arabic and not english:
        arabic = raw_name

    return arabic, english.lower()


def extract_company_code(raw_code):
    """
    Extract the numeric company code from the first column.

    Example: "450000112; يونايتد-united" -> "450000112"
    """
    if not raw_code:
        return ""
    match = re.match(r"(\d+)", str(raw_code).strip())
    return match.group(1) if match else ""


def process_excel():
    """Read the Excel file and build the knowledge base."""
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]
    print(f"  Loaded: {ws.max_row} rows x {ws.max_column} columns")

    # Build knowledge base
    knowledge_base = {}

    for row_idx in range(2, ws.max_row + 1):  # Skip header row
        raw_code = str(ws.cell(row_idx, 1).value or "")
        raw_company = str(ws.cell(row_idx, 2).value or "").strip()
        category = str(ws.cell(row_idx, 3).value or "").strip()
        details = clean_text(ws.cell(row_idx, 4).value)
        notes = clean_text(ws.cell(row_idx, 5).value)

        if not raw_company or not category:
            continue

        company_key = raw_company  # Use the full name as key

        if company_key not in knowledge_base:
            ar_name, en_name = extract_company_names(raw_company)
            company_id = extract_company_code(raw_code)

            knowledge_base[company_key] = {
                "company_id": company_id,
                "company_name": raw_company,
                "company_name_ar": ar_name,
                "company_name_en": en_name,
                "policies": {},
            }

        # Add the policy category
        category_en = CATEGORY_TRANSLATIONS.get(category, category)
        knowledge_base[company_key]["policies"][category] = {
            "category_ar": category,
            "category_en": category_en,
            "details": details,
            "notes": notes,
        }

    return knowledge_base


def main():
    print("=" * 60)
    print("PHASE 1: BUILDING INSURANCE KNOWLEDGE BASE")
    print("=" * 60)

    kb = process_excel()

    # Save to JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)

    # Summary statistics
    total_policies = sum(len(c["policies"]) for c in kb.values())
    categories_used = set()
    for company in kb.values():
        categories_used.update(company["policies"].keys())

    print(f"\n--- Knowledge Base Summary ---")
    print(f"  Total companies:  {len(kb)}")
    print(f"  Total policy entries: {total_policies}")
    print(f"  Unique categories: {len(categories_used)}")
    print(f"  Categories: {', '.join(sorted(categories_used))}")
    print(f"\n  Saved to: {OUTPUT_PATH}")

    # Show a few examples
    print(f"\n--- Sample Companies ---")
    for i, (key, company) in enumerate(kb.items()):
        if i >= 5:
            break
        print(f"  {company['company_name']} ({company['company_id']})")
        print(f"    AR: {company['company_name_ar']}")
        print(f"    EN: {company['company_name_en']}")
        print(f"    Categories: {len(company['policies'])}")

    print("\n[DONE] Knowledge base built successfully!")


if __name__ == "__main__":
    main()
